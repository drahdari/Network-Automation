import os
import sys
import xlsxwriter
from datetime import date
from openpyxl import workbook
from openpyxl.styles import Color, Font
from openpyxl.styles import Alignment
from openpyxl.styles.colors import GREEN
import openpyxl
import smtplib, ssl
import email.utils
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email import encoders
import xlrd
import csv

def Check_Emptiness(SString:str) -> bool:
    Empty=True
    for letter in SString:
        if letter!=" ":
            Empty=False
    return Empty
   

def Check_Lines(Line1:list, line2:list, check:str):
    try:
        line1_LIST = str(Line1).split(" ")
        line1_LIST_E=list()
        output=False
        reason="IP Difference"
        for items in line1_LIST:
            if Check_Emptiness(items) == False:
                line1_LIST_E.append(items)

        line2_LIST = str(line2).split(" ")
        line2_LIST_E = list()
        for items in line2_LIST:
            if Check_Emptiness(items) == False:
                line2_LIST_E.append(items)

        
        if len(line1_LIST_E) > 1 and len(line2_LIST_E) > 1:
            if str(line1_LIST_E[1]) == (line2_LIST_E[1]):
                    output = True
                    if check=="Added":
                        if line1_LIST_E[3] != line2_LIST_E[3]:
                            output = False
                            reason="MAC Changed"
                        if line1_LIST_E[5] != line2_LIST_E[5]:
                            output = False
                            reason="Vlan Changed"
        line1_LIST_E.append(output)
        line1_LIST_E.append(reason)
    except:
        print("Error")
        print(line1_LIST_E[1] + line2_LIST_E[1])

    return line1_LIST_E
                        

def Diff_Files(Folder1:str, Folder2:str, Check:str):    
    
    for f1 in os.listdir(Folder1):
        print(f1)
        for f2 in os.listdir(Folder2):
            
            if f1 == f2:
                TF= open(str(Folder1)+"\\" + str(f1), "r")
                for line1 in TF.readlines():
                    if len(str(line1)) > 1:
                        EXISTED= False
                        YF= open(str(Folder2) +"\\" + str(f2), "r")
                        
                        for line2 in YF.readlines():
                            line_LIST_E= list(Check_Lines(line1,line2, Check))
                            if str(line_LIST_E[7]).upper() == "MAC CHANGED":
                                    break
                            if bool(line_LIST_E[6]) == True:
                                EXISTED=True
                                break

                        if EXISTED == False:
                            F1_LIST = str(f1).split(".")
    
                            if os.path.isfile(str(OutputFile)):
                            
                                Create_EXCEl_FROM_TEXT(str(OutputFile),True, F1_LIST ,line_LIST_E, Check)
                            else:
                            
                                Create_EXCEl_FROM_TEXT(str(OutputFile),False, F1_LIST ,line_LIST_E, Check)
                
                        YF.close()
                    TF.close()



def Create_EXCEl_FROM_TEXT(FileName:str, Existed:bool, DeviceData:list, ARPData:list, state):

    SheetN=int()
    if str(ARPData[7]) == "IP Difference":
        if state=="Added":
            SheetN=0
        else:
            SheetN=1       
    elif str(ARPData[3])=="Incomplete":
        SheetN=2
    elif str(ARPData[7]) == "MAC Changed":
        SheetN=3
    else:
        SheetN=1

    if Existed:
        WB = openpyxl.load_workbook(str(FileName))
        Sheet_LIST = WB.get_sheet_names()
        sheet = WB.get_sheet_by_name(str(Sheet_LIST[SheetN]))
  
        Code_Site=str(Check_MACK(str(ARPData[3]),"Z:\\ARP_CACHE\\CodeSites"))

        str_row= [str(DeviceData[0]), str(DeviceData[1]), str(ARPData[1]), str(ARPData[3]), str(ARPData[5]), Code_Site]
        sheet.append(str_row)
        WB.save(str(FileName))
    else:
         WOrksheet_NAMES=["Added","Deleted","Incomplete","MAC Changed"]
         WB = openpyxl.Workbook() 
         c=0
         for name in WOrksheet_NAMES:
            if c!=0:
                WB.create_sheet(str(name))
            else:
                sheet=WB.active
                sheet.title = str(name)
            WB.active=c
            sheet=WB.active
            fnt = Font(size=11 , bold= True)
            """
            """
            sheet['A1'].font=fnt
            sheet['A1']='Device'
            sheet['A1'].alignment= Alignment(horizontal='center')
            sheet['B1']='Technology'
            sheet['B1'].alignment= Alignment(horizontal='center')
            sheet['B1'].font=fnt
            sheet['C1']='IP Address'
            sheet['C1'].alignment= Alignment(horizontal='center')
            sheet['C1'].font=fnt
            sheet['D1']='Mac Address'
            sheet['D1'].alignment= Alignment(horizontal='center')
            sheet['D1'].font=fnt
            sheet['E1'].alignment= Alignment(horizontal='center')
            sheet['E1']='Vlan Number'
            sheet['E1'].font=fnt
            sheet['F1'].alignment= Alignment(horizontal='center')
            sheet['F1']='Code Site'
            sheet['F1'].font=fnt
            c+=1
         Sheet_LIST = WB.get_sheet_names()
    
         Code_Site=str(Check_MACK(str(ARPData[3]),"Z:\\ARP_CACHE\\CodeSites"))
      
         str_row= [str(DeviceData[0]), str(DeviceData[1]), str(ARPData[1]), str(ARPData[3]), str(ARPData[5]), Code_Site]
         sheet = WB.get_sheet_by_name(str(Sheet_LIST[SheetN]))
      
         sheet.append(str_row)
         for dim in sheet.column_dimensions.values():
            dim.width=50
        
      
        
         WB.save(str(FileName))
      

def Check_MACK(MAC_Address:str, Folder:str) -> str:
    Found=False  
    Code_site="NOT EXISTED"
    for File_Path in os.listdir(Folder):
        if Found==False and str(File_Path).endswith(".xlsx"):                  
            workbook = xlrd.open_workbook(Folder+"\\"+File_Path)
            for sheet in workbook.sheets():
                ####Find Column#####
                CSite_Column=200
                MAC_Column=200
                cl=0
                while True:    
                    value= str(sheet.cell_value(0,cl))
                    if str(value).__contains__("MeContext_id"):
                        CSite_Column=cl
                    elif str(value).__contains__("mac"):
                        MAC_Column=cl 
                    if MAC_Column != 200 and CSite_Column !=200:
                        break            
                    cl+=1

                #####Find Match####                
                row=1
                EOF=False
                while(Found==False and not EOF):   
                    value=str(sheet.cell_value(row,MAC_Column))
                    if len(value)>0: 
                        MAC_Address=MAC_Address.replace(":","")
                        MAC_Address=MAC_Address.replace(".","")
                        value=value.replace(":","")
                        value=value.replace(".","")                
                        if value == MAC_Address:
                                Found=True
                                Code_site=str(sheet.cell_value(row,CSite_Column))
                        else:
                            if row < (sheet.nrows-1):
                                row += 1
                                
                            else:
                                EOF=True
                    else:
                        EOF=True
    return Code_site



def send_mail(file:str):
    password = "123d!@#D"
    context= ssl.SSLContext(ssl.PROTOCOL_SSLv23)
    msg = MIMEMultipart()
    msg['To'] = email.utils.formataddr(('IPRAN', 'danial.g@mtnirancell.ir'))
    msg['From'] = email.utils.formataddr(('ArpCheker', 'danial.r@mtnirancell.ir'))
    msg['Subject'] = "ARP Change of " + str(date.today()) 
    body = 'Dears' + "\n\n\n" + 'Please find the Latest ARP Change of Cisco Devices in the Attachment for ' + str(date.today()) + "\n\n\n\n" + 'Developed By MTN Irancell Operation Team' +"\n\n\n\n"
    msg.attach(MIMEText(body, 'plain'))
    
    server = smtplib.SMTP("mail.mtnirancell.ir", timeout=10, port=587)
    server.starttls(context=context)
    server.ehlo()
    server.login("danial.r@mtnirancell.ir", password)

    fp = open(file, 'rb')
    part = MIMEBase('application','vnd.ms-excel')
    part.set_payload(fp.read())
    fp.close()
    Base_File= file.split("\\")
    fName = Base_File[3]
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', 'attachment', filename="Cisco IPRAN Arp Statistics---" + fName)
    msg.attach(part)
    server.sendmail('danial.r@mtnirancell.ir', ['danial.r@mtnirancell.ir'], msg.as_string())
    server.quit()


def Folder_CUT(src: str, dst: str):
    for f1 in os.listdir(src):     
        if str(f1).endswith(".txt"):
            os.rename(src+"\\"+str(f1), dst +"\\"+str(f1))

def Folder_Remove(Folder: str):
    for f1 in os.listdir(Folder):
        os.remove(Folder+"\\"+f1)




Folder_Remove("Z:\ARP_CACHE\Yesterday")

Folder_CUT("Z:\ARP_CACHE\Today", "Z:\ARP_CACHE\Yesterday")

Folder_CUT("Z:\ARP_CACHE", "Z:\ARP_CACHE\Today")


TO_Files=list()
YE_Files=list()

OutputFile = "Z:\ARP_CACHE\Today" +"\\" + str(date.today())+".xlsx"
os.makedirs("Z:\ARP_CACHE\Today",exist_ok=True)


Diff_Files("Z:\ARP_CACHE\Today", "Z:\ARP_CACHE\Yesterday", "Added")
Diff_Files("Z:\ARP_CACHE\Yesterday", "Z:\ARP_CACHE\Today", "Removed")



if os.path.isfile(str(OutputFile)):
    send_mail(OutputFile)


    
                        
